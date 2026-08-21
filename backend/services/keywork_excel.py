"""重点工作 KeyWork Excel 模版生成与批量导入。

基于已部署的 PmwbKeyWork 数据模型（主表 + 7 张子表），提供：
1. build_template_bytes()  —— 生成「多页签分层」Excel 模版，供管理员离线填写。
2. import_key_works_from_bytes(db, raw) —— 解析上传的 xlsx，校验后原子入库。

模版结构（工作标识 work_key 作为主子表关联键，系统自动生成 KW- 编号）：
- 填写说明
- 重点工作（主表）
- 目标指标 / 里程碑 / 团队成员 / 月度计划 / 周计划 / 进展日志 / 成员待办（子表）
"""
from __future__ import annotations

import io
import json
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from db.models import (
    PmwbKeyWork,
    PmwbKeyWorkDeliverable,
    PmwbKeyWorkGoal,
    PmwbKeyWorkMember,
    PmwbKeyWorkMemberTask,
    PmwbKeyWorkMilestone,
    PmwbKeyWorkMonthlyPlan,
    PmwbKeyWorkProgress,
    PmwbKeyWorkWeeklyPlan,
)
from services.keywork import keywork_service

# ---------------------------------------------------------------------------
# 枚举取值（与 schemas/keywork.py 保持一致）
# ---------------------------------------------------------------------------
ENUM_OPTIONS: Dict[str, List[str]] = {
    "category": ["hq_pilot", "annual_task", "special_topic"],
    "priority": ["P0", "P1", "P2", "P3"],
    "status": ["planning", "in_progress", "completed", "paused", "cancelled"],
    "milestone_status": ["not_started", "in_progress", "completed", "cancelled", "delayed"],
    "plan_status": ["not_started", "in_progress", "completed", "cancelled", "delayed"],
    "task_status": ["not_started", "in_progress", "completed", "cancelled", "delayed"],
}

ENUM_LABELS: Dict[str, Dict[str, str]] = {
    "category": {"hq_pilot": "总部试点", "annual_task": "年度任务", "special_topic": "专题工作"},
    "priority": {"P0": "P0(最高)", "P1": "P1", "P2": "P2", "P3": "P3(最低)"},
    "status": {
        "planning": "规划中", "in_progress": "进行中", "completed": "已完成",
        "paused": "已暂停", "cancelled": "已取消",
    },
    "milestone_status": {
        "not_started": "未开始", "in_progress": "进行中", "completed": "已完成",
        "cancelled": "已作废", "delayed": "已延期",
    },
    "plan_status": {
        "not_started": "未开始", "in_progress": "进行中", "completed": "已完成",
        "cancelled": "已作废", "delayed": "已延期",
    },
    "task_status": {
        "not_started": "未开始", "in_progress": "进行中", "completed": "已完成",
        "cancelled": "已作废", "delayed": "已延期",
    },
}

DEFAULTS = {
    "category": "annual_task",
    "priority": "P2",
    "status": "planning",
    "milestone_status": "not_started",
    "plan_status": "not_started",
    "task_status": "not_started",
}

# ---------------------------------------------------------------------------
# 页签与列定义
# col: h=表头, k=key, enum=枚举名, date=日期列, req=必填, multi=多行
# ---------------------------------------------------------------------------
MAIN_SHEET = "重点工作"
CHILD_SHEETS = ["目标指标", "里程碑", "团队成员", "月度计划", "周计划", "进展日志", "成员待办"]

SHEETS: Dict[str, List[Dict[str, Any]]] = {
    MAIN_SHEET: [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "工作标题*", "k": "title", "req": True},
        {"h": "分类", "k": "category", "enum": "category"},
        {"h": "关联业务领域编码", "k": "domain_code"},
        {"h": "牵头人/负责人", "k": "owner"},
        {"h": "优先级", "k": "priority", "enum": "priority"},
        {"h": "状态", "k": "status", "enum": "status"},
        {"h": "进度百分比(0-100)", "k": "progress"},
        {"h": "计划完成时间(YYYY-MM-DD)", "k": "planned_finish_date", "date": True},
        {"h": "工作背景", "k": "background"},
        {"h": "现状说明", "k": "current_status"},
        {"h": "工作内容", "k": "content"},
        {"h": "验收标准(每行一条)", "k": "acceptance_criteria", "multi": True},
    ],
    "目标指标": [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "序号", "k": "seq"},
        {"h": "指标名称", "k": "indicator"},
        {"h": "目标值", "k": "target_value"},
        {"h": "当前值", "k": "current_value"},
        {"h": "单位", "k": "unit"},
        {"h": "说明", "k": "description"},
    ],
    "里程碑": [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "序号", "k": "seq"},
        {"h": "里程碑名称*", "k": "name", "req": True},
        {"h": "计划完成日期(YYYY-MM-DD)", "k": "due_date", "date": True},
        {"h": "状态", "k": "status", "enum": "milestone_status"},
        {"h": "说明", "k": "note"},
    ],
    "团队成员": [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "成员姓名*", "k": "name", "req": True},
        {"h": "角色", "k": "role"},
        {"h": "分工说明", "k": "division_desc"},
    ],
    "月度计划": [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "月份*(YYYY-MM)", "k": "month", "req": True},
        {"h": "创建日期(YYYY-MM-DD)", "k": "task_date", "date": True},
        {"h": "任务标题", "k": "title"},
        {"h": "任务描述", "k": "content"},
        {"h": "责任人", "k": "assignee"},
        {"h": "计划完成日期(YYYY-MM-DD)", "k": "due_date", "date": True},
        {"h": "状态", "k": "status", "enum": "plan_status"},
    ],
    "周计划": [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "周次*(YYYY-Www)", "k": "week", "req": True},
        {"h": "创建日期(YYYY-MM-DD)", "k": "task_date", "date": True},
        {"h": "任务标题", "k": "title"},
        {"h": "任务描述", "k": "content"},
        {"h": "责任人", "k": "assignee"},
        {"h": "计划完成日期(YYYY-MM-DD)", "k": "due_date", "date": True},
        {"h": "状态", "k": "status", "enum": "plan_status"},
    ],
    "进展日志": [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "进展日期(YYYY-MM-DD)", "k": "record_date", "date": True},
        {"h": "汇报人", "k": "reporter"},
        {"h": "进展内容", "k": "content"},
    ],
    "成员待办": [
        {"h": "工作标识*", "k": "work_key", "req": True},
        {"h": "待办标题*", "k": "title", "req": True},
        {"h": "负责人", "k": "assignee"},
        {"h": "截止日期(YYYY-MM-DD)", "k": "due_date", "date": True},
        {"h": "状态", "k": "status", "enum": "task_status"},
        {"h": "备注", "k": "note"},
    ],
}

# 子表 -> 模型 + 字段映射
CHILD_CONFIG = {
    "目标指标": (PmwbKeyWorkGoal, ["seq", "indicator", "target_value", "current_value", "unit", "description"]),
    "里程碑": (PmwbKeyWorkMilestone, ["seq", "name", "due_date", "status", "note"]),
    "团队成员": (PmwbKeyWorkMember, ["name", "role", "division_desc"]),
    "月度计划": (PmwbKeyWorkMonthlyPlan, ["month", "task_date", "title", "content", "assignee", "due_date", "status"]),
    "周计划": (PmwbKeyWorkWeeklyPlan, ["week", "task_date", "title", "content", "assignee", "due_date", "status"]),
    "进展日志": (PmwbKeyWorkProgress, ["record_date", "reporter", "content"]),
    "成员待办": (PmwbKeyWorkMemberTask, ["title", "assignee", "due_date", "status", "note"]),
}

# ---------------------------------------------------------------------------
# 样式
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="2F5496")
_SUB_FONT = Font(bold=True, size=11, color="2F5496")
_REQ_FONT = Font(color="C00000")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center")

_COL_WIDTHS = {
    MAIN_SHEET: [14, 24, 12, 18, 14, 10, 12, 14, 20, 28, 24, 28, 30],
    "目标指标": [14, 8, 18, 16, 16, 10, 30],
    "里程碑": [14, 8, 28, 22, 12, 30],
    "团队成员": [14, 14, 18, 36],
    "月度计划": [14, 16, 18, 24, 40, 14, 18, 12],
    "周计划": [14, 16, 18, 24, 40, 14, 18, 12],
    "进展日志": [14, 22, 14, 50],
    "成员待办": [14, 30, 14, 22, 12, 30],
}


# ---------------------------------------------------------------------------
# 模版生成
# ---------------------------------------------------------------------------
def build_template_bytes() -> io.BytesIO:
    """生成多页签 Excel 模版，返回 BytesIO。"""
    wb = Workbook()
    wb.remove(wb.active)

    _build_instruction_sheet(wb)
    for name in [MAIN_SHEET] + CHILD_SHEETS:
        _build_data_sheet(wb, name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_instruction_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("填写说明")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    r = 1
    ws.cell(r, 1, "重点工作 Excel 导入模版 · 填写说明").font = _TITLE_FONT
    r += 2

    lines = [
        ("一、总体说明", ""),
        ("", "本模版用于一次性收集多项『重点工作』及其全部明细，交给各管理员离线填写后，由系统【导入】功能批量入库。"),
        ("", "每一行 = 一条记录。各页签通过『工作标识』列关联：子表里的『工作标识』必须与『重点工作』页签中某行的『工作标识』完全一致，才会挂到对应重点工作下。"),
        ("二、工作标识（必填）", ""),
        ("", "在『重点工作』页签中由填写人自定义，例如 KW001、KW002。同一文件内必须唯一，仅用于关联，不会写入系统；系统入库时会自动生成 KW-YYYYMMDD-XXX 正式编号。"),
        ("三、各页签字段", ""),
        ("重点工作", "工作标识*、工作标题* 必填；分类/优先级/状态 从下拉选择；进度可填 0-100 整数或 0-1 小数（0.5 表示 50%）；计划完成时间可填 YYYY-MM-DD 或 YYYYMMDD 整数（如 20261030）；验收标准每行一条。"),
        ("目标指标", "逐条填写量化目标（指标名称/目标值/当前值/单位）。"),
        ("里程碑", "里程碑名称* 必填；状态从下拉选择（未开始/进行中/已完成/已延期）。"),
        ("团队成员", "成员姓名* 必填；可填角色与分工说明。"),
        ("月度计划 / 周计划", "月份填 YYYY-MM（如 2026-08），周次填 YYYY-Www（如 2026-W32）；创建日期、计划完成日期填 YYYY-MM-DD 或 YYYYMMDD；任务标题/任务描述/责任人均可填；状态从下拉选择（not_started/in_progress/completed/cancelled/delayed）"),
        ("进展日志", "记录工作进展，进展日期填 YYYY-MM-DD 或 YYYYMMDD，汇报人填姓名。"),
        ("成员待办", "待办标题* 必填；负责人填成员姓名；状态从下拉选择。"),
        ("四、日期格式", ""),
        ("", "所有日期支持以下格式：① YYYY-MM-DD（如 2026-08-31）② YYYYMMDD 整数（如 20261030）③ Excel 标准日期单元格。推荐直接输入文本日期，避免不同电脑日期格式差异。"),
        ("五、枚举取值对照", ""),
    ]
    for title, body in lines:
        if title and not body:
            ws.cell(r, 1, title).font = _SUB_FONT
            r += 1
        elif title:
            ws.cell(r, 1, title).font = Font(bold=True)
            ws.cell(r, 2, body).alignment = _WRAP
            r += 1
        else:
            ws.cell(r, 2, body).alignment = _WRAP
            r += 1

    # 枚举对照表
    r += 1
    ws.cell(r, 1, "字段").font = _HEADER_FONT
    ws.cell(r, 1).fill = _HEADER_FILL
    ws.cell(r, 2, "可选值（英文代码 = 中文含义）").font = _HEADER_FONT
    ws.cell(r, 2).fill = _HEADER_FILL
    r += 1
    enum_caption = {
        "category": "分类", "priority": "优先级", "status": "状态",
        "milestone_status": "里程碑状态", "plan_status": "计划/待办状态", "task_status": "成员待办状态",
    }
    for ek, caption in enum_caption.items():
        opts = " / ".join(f"{o} = {ENUM_LABELS[ek][o]}" for o in ENUM_OPTIONS[ek])
        ws.cell(r, 1, caption).font = Font(bold=True)
        ws.cell(r, 2, opts).alignment = _WRAP
        r += 1

    r += 1
    ws.cell(r, 1, "六、注意事项").font = _SUB_FONT
    r += 1
    for tip in [
        "带 * 的列为必填；留空行会被自动忽略，不会入库。",
        "请勿修改页签名称与各列表头文字，否则系统无法识别。",
        "重复导入同一文件会再次生成新记录（导入=新增，不做去重），请确认无误后导入。",
    ]:
        ws.cell(r, 2, "• " + tip).alignment = _WRAP
        r += 1


def _build_data_sheet(wb: Workbook, name: str) -> None:
    ws = wb.create_sheet(name)
    cols = SHEETS[name]
    widths = _COL_WIDTHS[name]

    # 表头
    for ci, col in enumerate(cols, start=1):
        c = ws.cell(1, ci, col["h"])
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
        if ci - 1 < len(widths):
            ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # 示例数据，降低填写门槛
    examples = {
        "重点工作": [
            ["KW001", "示例重点工作", "special_topic", "", "张三", "P1", "in_progress", 0.5, "2026-10-30", "工作背景示例", "现状说明示例", "工作内容示例", "验收标准示例"],
        ],
        "目标指标": [["KW001", 1, "指标示例", "1", "0.5", "百分比", ""]],
        "里程碑": [["KW001", 1, "里程碑示例", "2026-10-30", "in_progress", ""]],
        "团队成员": [["KW001", "张三", "SA", "总负责人"]],
        "月度计划": [["KW001", "2026-08", "2026-08-31", "月度任务示例", "任务描述", "张三", "2026-10-30", "not_started"]],
        "周计划": [["KW001", "2026-W32", "2026-08-19", "周任务示例", "任务描述", "张三", "2026-08-23", "not_started"]],
        "进展日志": [["KW001", "2026-08-19", "张三", "进展内容示例"]],
        "成员待办": [["KW001", "待办示例", "张三", "2026-08-23", "not_started", ""]],
    }
    if name in examples:
        for ri, row in enumerate(examples[name], start=2):
            for ci, val in enumerate(row, start=1):
                ws.cell(ri, ci, val)

    # 数据校验：枚举下拉 + 日期格式
    last_row = 1000
    enum_cols = [(ci + 1, col["enum"]) for ci, col in enumerate(cols) if col.get("enum")]
    date_cols = [(ci + 1) for ci, col in enumerate(cols) if col.get("date")]
    for col_idx, enum_name in enum_cols:
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(ENUM_OPTIONS[enum_name]) + '"',
            allow_blank=True,
        )
        dv.error = "请从下拉列表中选择有效值"
        dv.prompt = "可选值见『填写说明』页签枚举对照表"
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row}")
    for col_idx in date_cols:
        for rr in range(2, last_row + 1):
            ws.cell(rr, col_idx).number_format = "yyyy-mm-dd"


# ---------------------------------------------------------------------------
# 导入解析
# ---------------------------------------------------------------------------
def _clean(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _parse_date(v: Any) -> date:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    # Excel 可能把日期存为整数（如 20261030）或浮点数序列号
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # 8 位整数视为 YYYYMMDD
        if isinstance(v, int) and 19000101 <= v <= 21001231:
            return datetime.strptime(str(v), "%Y%m%d").date()
        # Excel 日期序列号（1 = 1900-01-01，2026-08-19 约 46253）
        if 1 <= v <= 100000:
            try:
                return datetime(1899, 12, 30) + timedelta(days=int(v))
            except (ValueError, OverflowError):
                pass

    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"日期格式应为 YYYY-MM-DD，收到：{s}")


def _parse_int(v: Any, field: str) -> int:
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        raise ValueError(f"{field} 应为整数，收到：{v}")


def _parse_progress(v: Any) -> int:
    """解析进度百分比：支持 0-1 小数（Excel 存储的 50% = 0.5）、0-100 整数、含 % 字符串。"""
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith("%"):
        try:
            return int(float(s[:-1]))
        except (ValueError, TypeError):
            raise ValueError(f"进度百分比格式错误，收到：{v}")
    try:
        f = float(s)
    except (ValueError, TypeError):
        raise ValueError(f"进度百分比应为数字，收到：{v}")
    # Excel 中 0.5 表示 50%，1 表示 100%
    if 0 <= f <= 1:
        return int(round(f * 100))
    return int(round(f))


def _read_sheet(ws, cols: List[Dict[str, Any]]) -> List[Tuple[int, Dict[str, Any]]]:
    """返回 (行号, 解析后的字段字典) 列表，跳过全空行。"""
    header_map: Dict[int, str] = {}
    for ci, col in enumerate(cols):
        cell_val = ws.cell(1, ci + 1).value
        if cell_val is not None and str(cell_val).strip() == col["h"]:
            header_map[ci] = col["k"]
    if not header_map:
        return []

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for r in range(2, ws.max_row + 1):
        vals: Dict[str, Any] = {}
        any_val = False
        for ci, key in header_map.items():
            raw = ws.cell(r, ci + 1).value
            v = _clean(raw)
            if v is not None:
                any_val = True
            vals[key] = v
        if not any_val:
            continue
        rows.append((r, vals))
    return rows


def import_key_works_from_bytes(db, raw: bytes) -> Dict[str, Any]:
    """解析 xlsx、校验、原子入库。返回 {ok, imported, total, errors}。"""
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return {"ok": False, "imported": 0, "total": 0,
                "errors": [{"sheet": "-", "row": 0, "message": f"无法读取 Excel 文件：{e}"}]}

    errors: List[Dict[str, Any]] = []
    parsed: Dict[str, Dict[str, Any]] = {}  # work_key -> {main, children:{sheet:[...]}}

    # ---- 主表 ----
    ws_main = wb.get_sheet_by_name(MAIN_SHEET)
    if ws_main is None:
        return {"ok": False, "imported": 0, "total": 0,
                "errors": [{"sheet": "-", "row": 0, "message": "缺少『重点工作』页签"}]}
    main_rows = _read_sheet(ws_main, SHEETS[MAIN_SHEET])
    for r, vals in main_rows:
        wk = vals.get("work_key")
        if not wk:
            errors.append({"sheet": MAIN_SHEET, "row": r, "message": "工作标识 必填"})
            continue
        if wk in parsed:
            errors.append({"sheet": MAIN_SHEET, "row": r, "message": f"工作标识『{wk}』重复"})
            continue
        # 必填
        if not vals.get("title"):
            errors.append({"sheet": MAIN_SHEET, "row": r, "message": "工作标题 必填"})
        # 枚举
        for col in SHEETS[MAIN_SHEET]:
            ek = col.get("enum")
            if ek and vals.get(col["k"]) is not None:
                if vals[col["k"]] not in ENUM_OPTIONS[ek]:
                    errors.append({"sheet": MAIN_SHEET, "row": r,
                                   "message": f"{col['h']} 取值非法：{vals[col['k']]}"})
        # 进度
        if vals.get("progress") is not None:
            try:
                p = _parse_progress(vals["progress"])
                if not (0 <= p <= 100):
                    errors.append({"sheet": MAIN_SHEET, "row": r, "message": "进度百分比需在 0-100 之间"})
            except ValueError as e:
                errors.append({"sheet": MAIN_SHEET, "row": r, "message": str(e)})
        # 日期
        if vals.get("planned_finish_date") is not None:
            try:
                _parse_date(vals["planned_finish_date"])
            except ValueError as e:
                errors.append({"sheet": MAIN_SHEET, "row": r, "message": str(e)})
        parsed[wk] = {"main": vals, "children": {s: [] for s in CHILD_SHEETS}}

    # ---- 子表 ----
    for sheet in CHILD_SHEETS:
        ws = wb.get_sheet_by_name(sheet)
        if ws is None:
            continue
        rows = _read_sheet(ws, SHEETS[sheet])
        for r, vals in rows:
            wk = vals.get("work_key")
            if not wk:
                errors.append({"sheet": sheet, "row": r, "message": "工作标识 必填"})
                continue
            if wk not in parsed:
                errors.append({"sheet": sheet, "row": r, "message": f"工作标识『{wk}』在『重点工作』页签中未找到"})
                continue
            # 必填子字段
            for col in SHEETS[sheet]:
                if col.get("req") and not vals.get(col["k"]):
                    errors.append({"sheet": sheet, "row": r, "message": f"{col['h']} 必填"})
            # 枚举
            for col in SHEETS[sheet]:
                ek = col.get("enum")
                if ek and vals.get(col["k"]) is not None:
                    if vals[col["k"]] not in ENUM_OPTIONS[ek]:
                        errors.append({"sheet": sheet, "row": r,
                                       "message": f"{col['h']} 取值非法：{vals[col['k']]}"})
            # 日期
            for col in SHEETS[sheet]:
                if col.get("date") and vals.get(col["k"]) is not None:
                    try:
                        _parse_date(vals[col["k"]])
                    except ValueError as e:
                        errors.append({"sheet": sheet, "row": r, "message": str(e)})
            parsed[wk]["children"][sheet].append((r, vals))

    total = len(parsed)
    if errors:
        return {"ok": False, "imported": 0, "total": total, "errors": errors}

    # ---- 校验通过，原子入库 ----
    try:
        created = []
        for wk, data in parsed.items():
            main = data["main"]
            obj_in = {
                "title": main["title"],
                "category": main.get("category") or DEFAULTS["category"],
                "domain_code": main.get("domain_code"),
                "owner": main.get("owner"),
                "priority": main.get("priority") or DEFAULTS["priority"],
                "status": main.get("status") or DEFAULTS["status"],
                "progress": _parse_progress(main.get("progress")) if main.get("progress") is not None else 0,
                "planned_finish_date": _parse_date(main["planned_finish_date"]) if main.get("planned_finish_date") else None,
                "background": main.get("background"),
                "current_status": main.get("current_status"),
                "content": main.get("content"),
                "acceptance_criteria": json.dumps(
                    [s.strip() for s in str(main.get("acceptance_criteria") or "").split("\n") if s.strip()],
                    ensure_ascii=False,
                ),
            }
            obj_in["work_no"] = keywork_service._gen_work_no(db)
            kw = PmwbKeyWork(**obj_in)
            db.add(kw)
            db.flush()

            for sheet, (model, fields) in CHILD_CONFIG.items():
                for _r, cvals in data["children"][sheet]:
                    child = {"key_work_id": kw.id}
                    for f in fields:
                        val = cvals.get(f)
                        if f in ("due_date", "record_date", "task_date"):
                            val = _parse_date(val) if val is not None else None
                        elif f == "seq":
                            val = _parse_int(val, "序号") if val is not None else None
                        elif f in ("status",):
                            val = val or DEFAULTS.get(_enum_for_status(sheet), "pending")
                        child[f] = val
                    db.add(model(**child))
            created.append(kw.id)
        db.commit()
        return {"ok": True, "imported": len(created), "total": total, "errors": []}
    except Exception as e:
        db.rollback()
        return {"ok": False, "imported": 0, "total": total,
                "errors": [{"sheet": "-", "row": 0, "message": f"入库失败：{e}"}]}


def _enum_for_status(sheet: str) -> str:
    if sheet == "里程碑":
        return "milestone_status"
    if sheet in ("月度计划", "周计划"):
        return "plan_status"
    if sheet == "成员待办":
        return "task_status"
    return "status"
