"""主动运营分析工单服务。

职责：
1. 生成优化后的分析工单 Excel 模板（双 sheet：填写区 + 填写说明）。
2. 解析导入的 Excel，创建分析工单（PmwbOperationIssue, category=prod）+ 分析明细
   （PmwbOperationAnalysis），并把「遗留任务」逐行自动建为「人员代办任务工单」
   （PmwbOperationIssue, category=task），责任人经人员中台解析，未匹配返回告警清单。
"""
from __future__ import annotations

import io
from datetime import datetime, date
from typing import Dict, List, Optional

import openpyxl
from sqlalchemy.orm import Session

from db.models import PmwbOperationIssue, PmwbOperationAnalysis
from services.operation import operation_issue_service
from services.staff_resolver import resolve_staff_id

# (字段key, 模板列A标签) —— 顺序即模板行顺序
TEMPLATE_FIELDS: List[tuple] = [
    ("topic_name", "课题名称"),
    ("analyst_team", "运营团队"),
    ("analyst_name", "运营人员"),
    ("domain_code", "关联业务领域编码"),
    ("background", "课题背景说明"),
    ("scenario", "操作场景介绍"),
    ("biz_flow", "业务流程梳理"),
    ("biz_rule", "业务规则梳理"),
    ("monitoring", "业务监控梳理"),
    ("analysis_goal", "本次分析目标"),
    ("data_analysis", "数据分析过程"),
    ("result_flow", "分析结果-流程优化方面"),
    ("result_rule", "分析结果-规则优化方面"),
    ("result_model", "分析结果-数据模型方面"),
    ("result_abnormal_user", "分析结果-异常用户数据方面"),
    ("result_monitor_blind", "分析结果-监控补盲方面"),
    ("go_live_date", "计划完成时间(yyyy-mm-dd)"),
]

LEGACY_HEADER = ["责任人", "任务内容", "计划完成时间", "优先级"]
VALID_PRIORITIES = ("P0", "P1", "P2", "P3")

_LABEL_TO_KEY = {label: key for key, label in TEMPLATE_FIELDS}


def _gen_no(prefix: str) -> str:
    return f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S%f')[:17]}"


def _parse_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def build_analysis_template_bytes() -> bytes:
    """生成主动运营分析工单模板，返回 xlsx 字节流。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分析工单"

    ws["A1"] = "主动运营分析工单填写模板"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    ws["A2"] = "填写说明见「填写说明」sheet；遗留任务填写在下方表格中，每行一条。"
    ws["A2"].font = openpyxl.styles.Font(italic=True, color="888888")

    start = 4
    for i, (key, label) in enumerate(TEMPLATE_FIELDS):
        r = start + i
        ws.cell(row=r, column=1, value=label).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=r, column=2, value="")

    # 遗留任务表：在字段区下方另起一段
    legacy_header_row = start + len(TEMPLATE_FIELDS) + 2
    ws.cell(row=legacy_header_row, column=1, value="遗留任务（未闭环任务登记，上传后自动建人员代办任务工单）").font = openpyxl.styles.Font(bold=True, size=12)
    hr = legacy_header_row + 1
    for c, h in enumerate(LEGACY_HEADER, start=1):
        ws.cell(row=hr, column=c, value=h).font = openpyxl.styles.Font(bold=True)
    # 预置 6 个空数据行
    for k in range(1, 7):
        rr = hr + k
        ws.cell(row=rr, column=1, value="")
        ws.cell(row=rr, column=2, value="")
        ws.cell(row=rr, column=3, value="")
        ws.cell(row=rr, column=4, value="P2")

    # 列宽
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10

    # 填写说明 sheet
    ws2 = wb.create_sheet("填写说明")
    notes = [
        "一、使用说明",
        "1. 在「分析工单」sheet 中按行填写分析内容，标 * 为重点字段。",
        "2. 课题名称为必填；其余分析字段尽量填全，便于沉淀与复盘。",
        "3. 「遗留任务」表格每行登记一条本周未闭环任务，上传后系统自动建为人员代办任务工单。",
        "",
        "二、字段字典",
        "运营团队 / 运营人员：本次课题的分析团队与负责人（分析工单的处理人取运营人员）。",
        "关联业务领域编码：可选，对应业务领域 domain_code。",
        "计划完成时间：yyyy-mm-dd 格式。",
        "优先级（遗留任务）：P0/P1/P2/P3，缺省 P2。",
        "",
        "三、自动同步说明",
        "上传模板后，系统会：① 创建一条「主动运营分析」工单（含分析明细）；",
        "② 把「遗留任务」每行生成一条「人员代办任务工单」，责任人为填写的姓名；",
        "③ 责任人姓名若在人员中台匹配不到，工单仍会创建但会返回未匹配清单，需人工认领。",
    ]
    for i, line in enumerate(notes, start=1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_analysis_fields(ws) -> Dict[str, str]:
    """按列A标签解析字段区。"""
    data: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        cells = list(row) + [None, None]
        a, b = cells[0], cells[1]
        if a is None:
            continue
        a = str(a).strip()
        if a in _LABEL_TO_KEY:
            val = b if b is not None else ""
            data[_LABEL_TO_KEY[a]] = str(val).strip() if val != "" else ""
    return data


def _parse_legacy_tasks(ws) -> List[Dict]:
    """定位「责任人」表头行，向下读取遗留任务数据。"""
    tasks: List[Dict] = []
    header_row = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a and str(a).strip() == "责任人":
            header_row = r
            break
    if not header_row:
        return tasks
    for r in range(header_row + 1, ws.max_row + 1):
        handler = ws.cell(row=r, column=1).value
        content = ws.cell(row=r, column=2).value
        if (not handler) and (not content):
            continue
        due = ws.cell(row=r, column=3).value
        priority = ws.cell(row=r, column=4).value
        tasks.append(
            {
                "handler": str(handler).strip() if handler else "",
                "content": str(content).strip() if content else "",
                "due_date": _parse_date(due),
                "priority": str(priority).strip() if priority else "P2",
            }
        )
    return tasks


def import_analysis_workbook(db: Session, file_bytes: bytes) -> Dict:
    """解析 Excel，落库分析工单 + 明细 + 遗留任务工单。整体在一个事务内。"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    # 取第一个 sheet 作为分析工单填写区
    ws = wb.worksheets[0]
    fields = _parse_analysis_fields(ws)
    legacy = _parse_legacy_tasks(ws)

    if not fields.get("topic_name"):
        raise ValueError("模板缺少「课题名称」，无法创建分析工单")

    issue_no = _gen_no("ANA")
    issue = PmwbOperationIssue(
        issue_no=issue_no,
        title=fields.get("topic_name", ""),
        category="prod",
        issue_type="topic_analysis",
        status="pending",
        source="analysis_import",
        handler=(fields.get("analyst_name") or fields.get("analyst_team") or ""),
        domain_code=fields.get("domain_code") or None,
        go_live_date=_parse_date(fields.get("go_live_date")),
    )
    db.add(issue)
    db.flush()  # 拿到 issue.id 用于外键

    detail = PmwbOperationAnalysis(
        issue_id=issue.id,
        topic_name=fields.get("topic_name") or None,
        analyst_team=fields.get("analyst_team") or None,
        analyst_name=fields.get("analyst_name") or None,
        domain_code=fields.get("domain_code") or None,
        background=fields.get("background") or None,
        scenario=fields.get("scenario") or None,
        biz_flow=fields.get("biz_flow") or None,
        biz_rule=fields.get("biz_rule") or None,
        monitoring=fields.get("monitoring") or None,
        analysis_goal=fields.get("analysis_goal") or None,
        data_analysis=fields.get("data_analysis") or None,
        result_flow=fields.get("result_flow") or None,
        result_rule=fields.get("result_rule") or None,
        result_model=fields.get("result_model") or None,
        result_abnormal_user=fields.get("result_abnormal_user") or None,
        result_monitor_blind=fields.get("result_monitor_blind") or None,
    )
    db.add(detail)

    unmatched: List[str] = []
    created_tasks = 0
    for t in legacy:
        if not t["content"]:
            continue
        handler = t["handler"]
        if handler:
            sid = resolve_staff_id(handler)
            if sid is None:
                unmatched.append(handler)
        priority = t["priority"] if t["priority"] in VALID_PRIORITIES else "P2"
        task = PmwbOperationIssue(
            issue_no=_gen_no("TASK"),
            title=t["content"],
            category="task",
            issue_type="temp_task",
            status="pending",
            source="analysis_legacy",
            handler=handler or "",
            impact_level=priority,
            go_live_date=t["due_date"],
            related_req_id=issue_no,
        )
        db.add(task)
        created_tasks += 1

    db.commit()
    db.refresh(issue)
    return {
        "issue_no": issue_no,
        "analysis_id": detail.id,
        "topic_name": fields.get("topic_name"),
        "legacy_task_count": created_tasks,
        "unmatched_handlers": unmatched,
    }


def get_analysis_detail(db: Session, issue_id: int) -> Dict:
    """取分析工单明细 + 关联遗留任务工单，供前端详情展示。"""
    issue = operation_issue_service.get(db, issue_id)
    if not issue:
        return {"issue": None, "analysis": None, "legacy_tasks": []}
    detail = (
        db.query(PmwbOperationAnalysis)
        .filter(PmwbOperationAnalysis.issue_id == issue_id)
        .first()
    )
    legacy_tasks = (
        db.query(PmwbOperationIssue)
        .filter(
            PmwbOperationIssue.related_req_id == issue.issue_no,
            PmwbOperationIssue.category == "task",
        )
        .order_by(PmwbOperationIssue.id)
        .all()
    )
    return {"issue": issue, "analysis": detail, "legacy_tasks": legacy_tasks}
