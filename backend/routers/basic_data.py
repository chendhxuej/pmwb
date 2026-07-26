"""基础数据路由：组织 + 人员主数据 CRUD、选人分组选项。

挂在 /api/v1/basic-data 下，是全站选人组件的统一数据源。
"""
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from core.exceptions import NotFoundException, ValidationException
from core.response import success
from db.base import get_db
from schemas.basic_data import (
    OrgCreate,
    OrgOut,
    OrgStaffImportOut,
    OrgUpdate,
    StaffCreate,
    StaffOptionGroup,
    StaffOut,
    StaffUpdate,
)
from services.basic_data import TEMPLATE_COLUMNS, basic_data_service

router = APIRouter(prefix="/basic-data", tags=["基础数据"])


# ---------------------------------------------------------------------------
# 选人组件分组选项（放在最前，避免与 /orgs/{id} 混淆）
# ---------------------------------------------------------------------------
@router.get("/staff-options")
def staff_options(db: Session = Depends(get_db)):
    """按组织分组返回全量启用人员选项（选人组件数据源）。"""
    groups = basic_data_service.staff_options(db)
    return success(data=[StaffOptionGroup(**g).model_dump() for g in groups])


# ---------------------------------------------------------------------------
# 组织 CRUD
# ---------------------------------------------------------------------------
@router.get("/orgs")
def list_orgs(db: Session = Depends(get_db)):
    orgs = basic_data_service.list_orgs(db)
    result = []
    for o in orgs:
        out = OrgOut.model_validate(o)
        out.staff_count = len(o.staffs)
        result.append(out.model_dump())
    return success(data=result)


@router.post("/orgs")
def create_org(obj_in: OrgCreate, db: Session = Depends(get_db)):
    obj = basic_data_service.create_org(db, obj_in.model_dump(exclude_unset=True))
    return success(data=OrgOut.model_validate(obj).model_dump(), message="组织已创建")


@router.put("/orgs/{org_id}")
def update_org(org_id: int, obj_in: OrgUpdate, db: Session = Depends(get_db)):
    obj = basic_data_service.update_org(db, org_id, obj_in.model_dump(exclude_unset=True))
    if not obj:
        raise NotFoundException(f"组织不存在：id={org_id}")
    return success(data=OrgOut.model_validate(obj).model_dump(), message="组织已更新")


@router.delete("/orgs/{org_id}")
def delete_org(org_id: int, db: Session = Depends(get_db)):
    ok = basic_data_service.delete_org(db, org_id)
    if not ok:
        raise NotFoundException(f"组织不存在：id={org_id}")
    return success(data={"deleted": True}, message="组织已删除")


# ---------------------------------------------------------------------------
# 人员 CRUD
# ---------------------------------------------------------------------------
@router.get("/staffs")
def list_staffs(
    org_id: Optional[int] = Query(None, description="按组织过滤"),
    keyword: Optional[str] = Query(None, description="姓名/邮箱关键字"),
    db: Session = Depends(get_db),
):
    staffs = basic_data_service.list_staffs(db, org_id=org_id, keyword=keyword)
    result = []
    for s in staffs:
        out = StaffOut.model_validate(s)
        out.org_name = s.org.name if s.org else None
        result.append(out.model_dump())
    return success(data=result)


@router.post("/staffs")
def create_staff(obj_in: StaffCreate, db: Session = Depends(get_db)):
    obj = basic_data_service.create_staff(db, obj_in.model_dump(exclude_unset=True))
    out = StaffOut.model_validate(obj)
    out.org_name = obj.org.name if obj.org else None
    return success(data=out.model_dump(), message="人员已创建")


@router.put("/staffs/{staff_id}")
def update_staff(staff_id: int, obj_in: StaffUpdate, db: Session = Depends(get_db)):
    obj = basic_data_service.update_staff(db, staff_id, obj_in.model_dump(exclude_unset=True))
    if not obj:
        raise NotFoundException(f"人员不存在：id={staff_id}")
    out = StaffOut.model_validate(obj)
    out.org_name = obj.org.name if obj.org else None
    return success(data=out.model_dump(), message="人员已更新")


@router.delete("/staffs/{staff_id}")
def delete_staff(staff_id: int, db: Session = Depends(get_db)):
    ok = basic_data_service.delete_staff(db, staff_id)
    if not ok:
        raise NotFoundException(f"人员不存在：id={staff_id}")
    return success(data={"deleted": True}, message="人员已删除")


# ---------------------------------------------------------------------------
# 批量导入与模板
# ---------------------------------------------------------------------------
@router.post("/import")
def import_basic_data(
    file: UploadFile = File(..., description="Excel 文件，列：组织名称, 成员姓名, 邮箱, 电话, 身份, 排序号, 是否启用"),
    db: Session = Depends(get_db),
):
    """上传 Excel 批量导入/更新组织与人员。"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise ValidationException("仅支持 .xlsx / .xls 文件")

    contents = file.file.read()
    try:
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as e:
        raise ValidationException(f"无法解析 Excel：{e}")

    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    expected = [c[0] for c in TEMPLATE_COLUMNS]
    # 允许缺少非必填列，但必填列必须存在
    required_names = {c[0] for c in TEMPLATE_COLUMNS if c[1]}
    missing_required = required_names - set(headers)
    if missing_required:
        raise ValidationException(f"Excel 缺少必填列：{', '.join(missing_required)}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if header in expected:
                row_dict[header] = row[idx] if idx < len(row) else None
        rows.append(row_dict)

    if not rows:
        raise ValidationException("Excel 中没有有效数据行")

    result = basic_data_service.import_orgs_staffs(db, rows)
    return success(
        data=OrgStaffImportOut(**result).model_dump(),
        message=f"导入完成：新增 {result['created_orgs']} 个组织、{result['created_staffs']} 名成员，更新 {result['updated_orgs']} 个组织、{result['updated_staffs']} 名成员",
    )


@router.get("/template")
def download_template():
    """下载团队信息导入 Excel 模板。"""
    data = basic_data_service.build_template_bytes()
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="team_info_template.xlsx"',
        },
    )
