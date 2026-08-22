"""人员主数据服务：组织 + 人员 CRUD、选人选项、批量导入。"""
from __future__ import annotations

import io
from collections import defaultdict
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from core.exceptions import ValidationException
from db.models import PmwbOrg, PmwbRole, PmwbStaff

# Excel 导入模板列定义
TEMPLATE_COLUMNS = [
    ("组织名称", True),
    ("成员姓名", True),
    ("邮箱", False),
    ("电话", False),
    ("身份", False),
    ("排序号", False),
    ("是否启用", False),
]


def _cell_str(value) -> str:
    """把 openpyxl 单元格值转字符串；空值/None 返回空字符串。"""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    return str(value).strip()


def _parse_enabled(value):
    """把文本/数字解析为布尔值。"""
    if value is None:
        return True
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("1", "true", "是", "yes", "y", "启用"):
        return True
    if s in ("0", "false", "否", "no", "n", "停用", "禁用"):
        return False
    return True


class BasicDataService:
    """组织+人员主数据服务。"""

    # ------------------------------------------------------------------
    # 组织
    # ------------------------------------------------------------------
    def list_orgs(self, db: Session, include_disabled: bool = True) -> List[PmwbOrg]:
        query = db.query(PmwbOrg)
        if not include_disabled:
            query = query.filter(PmwbOrg.enabled.is_(True))
        return query.order_by(PmwbOrg.sort, PmwbOrg.id).all()

    def create_org(self, db: Session, data: dict) -> PmwbOrg:
        from sqlalchemy.exc import IntegrityError

        obj = PmwbOrg(**data)
        db.add(obj)
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            raise ValidationException(f"组织名称重复：{data.get('name', '')}")
        db.refresh(obj)
        return obj

    def update_org(self, db: Session, org_id: int, data: dict) -> Optional[PmwbOrg]:
        obj = db.get(PmwbOrg, org_id)
        if not obj:
            return None
        for k, v in data.items():
            setattr(obj, k, v)
        db.commit()
        db.refresh(obj)
        return obj

    def delete_org(self, db: Session, org_id: int) -> bool:
        obj = db.get(PmwbOrg, org_id)
        if not obj:
            return False
        # 级联删除该组织下人员，避免删除组织后留下孤儿人员数据
        db.query(PmwbStaff).filter(PmwbStaff.org_id == org_id).delete()
        db.delete(obj)
        db.commit()
        return True

    # ------------------------------------------------------------------
    # 角色/身份定义
    # ------------------------------------------------------------------
    def list_roles(self, db: Session, include_disabled: bool = True) -> List[PmwbRole]:
        query = db.query(PmwbRole)
        if not include_disabled:
            query = query.filter(PmwbRole.enabled.is_(True))
        return query.order_by(PmwbRole.sort, PmwbRole.id).all()

    def create_role(self, db: Session, data: dict) -> PmwbRole:
        obj = PmwbRole(**data)
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    def update_role(self, db: Session, role_id: int, data: dict) -> Optional[PmwbRole]:
        obj = db.get(PmwbRole, role_id)
        if not obj:
            return None
        for k, v in data.items():
            setattr(obj, k, v)
        db.commit()
        db.refresh(obj)
        return obj

    def delete_role(self, db: Session, role_id: int) -> bool:
        obj = db.get(PmwbRole, role_id)
        if not obj:
            return False
        db.delete(obj)
        db.commit()
        return True

    # ------------------------------------------------------------------
    # 轻量选项（下拉用，不加载人员明细）
    # ------------------------------------------------------------------
    def org_options(self, db: Session) -> List[dict]:
        """返回启用的组织名称列表（用于选人组件的组织下拉）。"""
        orgs = (
            db.query(PmwbOrg)
            .filter(PmwbOrg.enabled.is_(True))
            .order_by(PmwbOrg.sort, PmwbOrg.id)
            .all()
        )
        return [{"id": o.id, "name": o.name} for o in orgs]

    def role_options(self, db: Session) -> List[dict]:
        """返回启用的身份名称列表（用于选人组件的身份下拉）。"""
        roles = (
            db.query(PmwbRole)
            .filter(PmwbRole.enabled.is_(True))
            .order_by(PmwbRole.sort, PmwbRole.id)
            .all()
        )
        return [{"id": r.id, "name": r.name} for r in roles]

    # ------------------------------------------------------------------
    # 人员
    # ------------------------------------------------------------------
    def list_staffs(
        self,
        db: Session,
        org_id: Optional[int] = None,
        keyword: Optional[str] = None,
        include_disabled: bool = True,
    ) -> List[PmwbStaff]:
        query = db.query(PmwbStaff)
        if org_id:
            query = query.filter(PmwbStaff.org_id == org_id)
        if keyword:
            like = f"%{keyword}%"
            query = query.filter(
                PmwbStaff.name.like(like)
                | PmwbStaff.email.like(like)
                | PmwbStaff.role_hint.like(like)
            )
        if not include_disabled:
            query = query.filter(PmwbStaff.enabled.is_(True))
        return query.order_by(PmwbStaff.org_id, PmwbStaff.sort, PmwbStaff.id).all()

    def create_staff(self, db: Session, data: dict) -> PmwbStaff:
        from sqlalchemy.exc import IntegrityError

        obj = PmwbStaff(**data)
        db.add(obj)
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            raise ValidationException(f"人员已存在：{data.get('name', '')}")
        db.refresh(obj)
        return obj

    def update_staff(self, db: Session, staff_id: int, data: dict) -> Optional[PmwbStaff]:
        obj = db.get(PmwbStaff, staff_id)
        if not obj:
            return None
        for k, v in data.items():
            setattr(obj, k, v)
        db.commit()
        db.refresh(obj)
        return obj

    def delete_staff(self, db: Session, staff_id: int) -> bool:
        obj = db.get(PmwbStaff, staff_id)
        if not obj:
            return False
        db.delete(obj)
        db.commit()
        return True

    # ------------------------------------------------------------------
    # 选人组件
    # ------------------------------------------------------------------
    def staff_options(self, db: Session) -> List[dict]:
        orgs = (
            db.query(PmwbOrg)
            .filter(PmwbOrg.enabled.is_(True))
            .order_by(PmwbOrg.sort, PmwbOrg.id)
            .all()
        )
        groups = []
        for org in orgs:
            options = [
                {
                    "value": s.name,
                    "label": s.name,
                    "email": s.email,
                    "role_hint": s.role_hint,
                }
                for s in sorted(org.staffs, key=lambda x: ((x.sort or 0), x.id))
                if s.enabled
            ]
            if options:
                groups.append({"org_id": org.id, "org_name": org.name, "options": options})
        return groups

    # ------------------------------------------------------------------
    # 批量导入
    # ------------------------------------------------------------------
    def import_orgs_staffs(self, db: Session, rows: List[dict]) -> dict:
        """从 Excel 解析后的行批量 upsert 组织与人员。"""
        stats = {
            "created_orgs": 0,
            "updated_orgs": 0,
            "created_staffs": 0,
            "updated_staffs": 0,
            "errors": [],
        }

        # 1. 按组织名称聚类
        org_rows: Dict[str, List[dict]] = defaultdict(list)
        for idx, row in enumerate(rows, start=2):
            org_name = _cell_str(row.get("组织名称", ""))
            if not org_name:
                stats["errors"].append(f"第 {idx} 行：组织名称不能为空")
                continue
            org_rows[org_name].append({"idx": idx, "row": row})

        # 2. 组织 upsert
        org_cache: Dict[str, PmwbOrg] = {}
        for org_name, entries in org_rows.items():
            representative = entries[0]["row"]
            org = db.query(PmwbOrg).filter(PmwbOrg.name == org_name).first()
            sort = int(_cell_str(representative.get("排序号", "")) or 0)
            enabled = _parse_enabled(representative.get("是否启用", True))
            if org:
                org.sort = sort
                org.enabled = enabled
                stats["updated_orgs"] += 1
            else:
                org = PmwbOrg(name=org_name, sort=sort, enabled=enabled, source_trace="manual")
                db.add(org)
                db.flush()
                stats["created_orgs"] += 1
            org_cache[org_name] = org

        # 3. 人员 upsert
        for org_name, entries in org_rows.items():
            org = org_cache[org_name]
            for item in entries:
                idx = item["idx"]
                row = item["row"]
                name = _cell_str(row.get("成员姓名", ""))
                if not name:
                    continue

                email = _cell_str(row.get("邮箱", "")) or None
                phone = _cell_str(row.get("电话", "")) or None
                role_hint = _cell_str(row.get("身份", "")) or None
                sort = int(_cell_str(row.get("排序号", "")) or 0)
                enabled = _parse_enabled(row.get("是否启用", True))

                staff = (
                    db.query(PmwbStaff)
                    .filter(PmwbStaff.name == name, PmwbStaff.org_id == org.id)
                    .first()
                )
                if staff:
                    staff.email = email
                    staff.phone = phone
                    staff.role_hint = role_hint
                    staff.sort = sort
                    staff.enabled = enabled
                    stats["updated_staffs"] += 1
                else:
                    staff = PmwbStaff(
                        name=name,
                        org_id=org.id,
                        email=email,
                        phone=phone,
                        role_hint=role_hint,
                        sort=sort,
                        enabled=enabled,
                        source_trace="manual",
                    )
                    db.add(staff)
                    stats["created_staffs"] += 1

        db.commit()
        return stats

    def build_template_bytes(self) -> bytes:
        """生成 Excel 导入模板字节流（含填写说明 + 数据页签）。"""
        wb = Workbook()
        wb.remove(wb.active)

        # 填写说明
        ws_help = wb.create_sheet("填写说明")
        ws_help.sheet_view.showGridLines = False
        ws_help.column_dimensions["A"].width = 18
        ws_help.column_dimensions["B"].width = 80

        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        sub_font = Font(bold=True, size=11, color="2F5496")
        wrap = Alignment(vertical="top", wrap_text=True)

        r = 1
        ws_help.cell(r, 1, "团队信息 Excel 导入模板 · 填写说明").font = Font(bold=True, size=14, color="2F5496")
        r += 2

        help_lines = [
            ("一、总体说明", ""),
            ("", "每一行 = 一名成员；相同组织+姓名的成员会被覆盖更新，不会重复创建。"),
            ("", "去重规则：优先按邮箱匹配；邮箱为空时按 组织名称+成员姓名 匹配。"),
            ("二、必填与选填", ""),
            ("组织名称*", "必填。若不存在则自动创建新组织。"),
            ("成员姓名*", "必填。"),
            ("邮箱", "可选但强烈建议填写，用于邮件中心、会议通知等场景。"),
            ("电话", "可选。"),
            ("身份", "可选，如产品经理、SA、开发等，用于角色筛选。"),
            ("排序号", "可选，数字越小排序越靠前，默认 0。"),
            ("是否启用", "可选，填“是/启用/1/true”表示启用，填“否/停用/0/false”表示停用，默认启用。"),
            ("三、注意事项", ""),
            ("", "• 请勿修改页签名称与各列表头文字。"),
            ("", "• 留空行会被自动忽略。"),
            ("", "• 导入前建议在人员中台页面查看现有人员，避免误覆盖。"),
        ]
        for title, body in help_lines:
            if title and not body:
                ws_help.cell(r, 1, title).font = sub_font
            elif title:
                ws_help.cell(r, 1, title).font = Font(bold=True)
                ws_help.cell(r, 2, body).alignment = wrap
            else:
                ws_help.cell(r, 2, body).alignment = wrap
            r += 1

        # 数据页签
        ws = wb.create_sheet("团队信息导入")
        headers = [c[0] for c in TEMPLATE_COLUMNS]
        ws.append(headers)
        for col_idx, (name, _required) in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = max(14, len(name) + 4)

        # 示例数据
        ws.append(["政企业务部", "张三", "zhangsan@example.com", "13800138000", "产品经理", 0, "是"])
        ws.append(["CRM维护", "李四", "lisi@example.com", "", "系统维护", 1, "是"])
        ws.append(["BOSS维护", "", "", "", "", "", ""])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()


basic_data_service = BasicDataService()
